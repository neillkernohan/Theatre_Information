#!/usr/bin/env python3
"""
Update "Note to House Manager" for patrons from the People by Role spreadsheet.
For each row:
  1. Look up patron by email in ArtsPeople
  2. If not found, create the patron record
  3. Add to the "Volunteers" mailing list (if not already on it)
  4. Append the Show - Role to the "Note to House Manager" field
  5. Update and exit the record
"""

import os
import time
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)
from selenium.webdriver.common.keys import Keys
from dotenv import load_dotenv

# Load env from same directory as this script
script_dir = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(script_dir, ".env"))

NEON_EMAIL = os.getenv("NEON_EMAIL")
NEON_PASSWORD = os.getenv("NEON_PASSWORD")

XLSX_FILE = os.path.join(script_dir, "Theatre_Aurora_People_By_Role v4.xlsx")
SHEET_NAME = "People by Role"

# ArtsPeople People page
PEOPLE_URL = "https://app.arts-people.com/admin/legacyurl/1090"

# Note to House Manager textarea
NOTE_FIELD_ID = "TXT109000026"

# Column to track which rows have been processed (added to spreadsheet)
PROCESSED_COL = "Processed"


def timestamp():
    return datetime.now().strftime("%H:%M:%S")


def log(icon, message):
    print(f"  {icon}  [{timestamp()}] {message}")


def login(driver):
    """Log into ArtsPeople via Neon SSO."""
    log("🔑", "Logging into Neon SSO...")
    driver.get("https://app.neonsso.com/login")

    email_input = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.ID, "email"))
    )
    email_input.send_keys(NEON_EMAIL)

    driver.find_element(By.XPATH, "//button[normalize-space()='Next']").click()

    password_input = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.ID, "password"))
    )
    password_input.send_keys(NEON_PASSWORD)

    driver.find_element(By.XPATH, "//button[normalize-space()='Log In']").click()

    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located(
            (By.XPATH, "//a[normalize-space()='Open']")
        )
    ).click()

    WebDriverWait(driver, 30).until(
        EC.url_contains("dashboard")
    )
    log("✅", "Logged in successfully.")

    navigate_to_people_page(driver)


def navigate_to_people_page(driver):
    """Navigate to the Database > People page."""
    driver.get(PEOPLE_URL)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "BTNlookupExisting"))
    )


def clear_screen(driver):
    """Click Clear Screen to reset the form."""
    try:
        clear_btn = driver.find_element(By.ID, "BTNpatron_clear_btn")
        clear_btn.click()
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "BTNlookupExisting"))
        )
        time.sleep(0.5)
    except Exception:
        navigate_to_people_page(driver)


def lookup_by_email(driver, email):
    """
    Search for a patron by email.
    Returns True if found (patron record loaded), False if not found.
    """
    email_field = driver.find_element(By.ID, "TXT3293")
    email_field.clear()
    email_field.send_keys(email)

    driver.find_element(By.ID, "BTNlookupExisting").click()

    time.sleep(2)

    # Check for "Select Customer" page (multiple matches)
    try:
        driver.find_element(By.XPATH, "//*[contains(text(),'Select Customer')]")
        return "multiple"
    except NoSuchElementException:
        pass

    try:
        driver.find_element(By.XPATH, "//*[contains(text(),'No customers found')]")
        return False
    except NoSuchElementException:
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//input[@value='Exit Record']")
                )
            )
            return True
        except TimeoutException:
            return False


def create_patron(driver, first_name, last_name, email):
    """Create a new patron with the given details."""
    first_field = driver.find_element(By.ID, "TXTfirst_name")
    first_field.clear()
    first_field.send_keys(first_name)

    last_field = driver.find_element(By.ID, "TXTlast_name")
    last_field.clear()
    last_field.send_keys(last_name)

    driver.find_element(By.ID, "BTNsave_patron").click()

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(
            (By.XPATH, "//input[@value='Exit Record']")
        )
    )
    time.sleep(1)


def is_on_volunteers_list(driver):
    """Check if the current patron is already on the Volunteers list."""
    try:
        elements = driver.find_elements(
            By.XPATH,
            "//h2[contains(text(),'Lists')]/following-sibling::*"
            "//*[normalize-space()='Volunteers'] | "
            "//*[contains(text(),'Lists')]"
            "/following::*[normalize-space()='Volunteers']"
        )
        for el in elements:
            try:
                if "Volunteers" in el.text:
                    return True
            except StaleElementReferenceException:
                continue
        return False
    except Exception:
        return False


def add_to_volunteers_list(driver):
    """
    Add the current patron to the Volunteers list.
    Assumes we're on a patron record page.
    Returns True if successful, False otherwise.
    """
    try:
        driver.execute_script("window.scrollTo(0, 500);")
        time.sleep(0.5)

        # Click the "Add to List" button
        add_btn = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, "//button[normalize-space()='Add to List']")
            )
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", add_btn)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", add_btn)
        time.sleep(1)

        # Wait for the slide-out panel
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(),'Add patron to lists')]")
            )
        )

        # Find and click "Volunteers" in the list
        volunteers_item = None
        try:
            volunteers_item = driver.find_element(
                By.XPATH,
                "//div[contains(@class,'add-patron-to-lists')]"
                "//span[normalize-space()='Volunteers'] | "
                "//div[contains(@class,'add-patron-to-lists')]"
                "//*[normalize-space()='Volunteers']"
            )
        except NoSuchElementException:
            try:
                volunteers_item = driver.find_element(
                    By.XPATH,
                    "//*[contains(text(),'Add patron to lists')]"
                    "/following::*[normalize-space()='Volunteers']"
                )
            except NoSuchElementException:
                pass

        if volunteers_item is None:
            log("ℹ️", "  'Volunteers' not in add list (may already be on it)")
            driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
            time.sleep(0.5)
            return True

        volunteers_item.click()
        time.sleep(0.5)

        # Click the confirm "Add to List" button in the panel
        panel_add_btn = driver.find_element(
            By.CSS_SELECTOR,
            "button.mailing-list-add-to-list:not(.ng-scope)"
        )
        driver.execute_script("arguments[0].click();", panel_add_btn)
        time.sleep(1)

        return True

    except Exception as e:
        log("❌", f"  Error adding to Volunteers list: {e}")
        return False


def append_note_to_house_manager(driver, role_text):
    """
    Append role_text to the Note to House Manager field.
    If the field already contains text, add on a new line.
    """
    note_field = driver.find_element(By.ID, NOTE_FIELD_ID)
    driver.execute_script("arguments[0].scrollIntoView(true);", note_field)
    time.sleep(0.3)

    existing = note_field.get_attribute("value") or ""

    # Check if this exact role is already in the note
    if role_text in existing:
        log("ℹ️", f"  Role already in note: {role_text}")
        return False

    if existing.strip():
        new_value = existing.rstrip() + "\n" + role_text
    else:
        new_value = role_text

    note_field.clear()
    note_field.send_keys(new_value)
    time.sleep(0.3)
    return True


def update_patron_record(driver):
    """Click 'Update Patron Record' to save changes."""
    try:
        update_btn = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, "//input[@value='Update Patron Record']")
            )
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", update_btn)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", update_btn)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, "//input[@value='Exit Record']")
            )
        )
        time.sleep(1)
        log("✅", "  Patron record updated")
    except Exception as e:
        log("⚠️", f"  Could not update patron record: {e}")


def exit_record(driver):
    """Click Exit Record to go back to the search page."""
    try:
        exit_btn = driver.find_element(
            By.XPATH, "//input[@value='Exit Record']"
        )
        driver.execute_script("arguments[0].click();", exit_btn)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "BTNlookupExisting"))
        )
        time.sleep(0.5)
    except Exception:
        navigate_to_people_page(driver)


def read_xlsx():
    """Read the People by Role sheet. Returns (workbook, rows_data)."""
    wb = load_workbook(XLSX_FILE)
    ws = wb[SHEET_NAME]

    # Check if Processed column exists, add it if not
    headers = [cell.value for cell in ws[1]]
    if PROCESSED_COL not in headers:
        proc_col_idx = len(headers) + 1
        ws.cell(row=1, column=proc_col_idx, value=PROCESSED_COL)
    else:
        proc_col_idx = headers.index(PROCESSED_COL) + 1

    rows = []
    for row_num in range(2, ws.max_row + 1):
        name = ws.cell(row=row_num, column=1).value or ""
        show_role = ws.cell(row=row_num, column=2).value or ""
        email = ws.cell(row=row_num, column=3).value or ""
        processed = ws.cell(row=row_num, column=proc_col_idx).value or ""

        rows.append({
            "row_num": row_num,
            "name": str(name).strip(),
            "show_role": str(show_role).strip(),
            "email": str(email).strip(),
            "processed": str(processed).strip(),
            "proc_col_idx": proc_col_idx,
        })

    return wb, ws, rows


def save_xlsx(wb):
    """Save the workbook."""
    wb.save(XLSX_FILE)


def split_name(name):
    """Split a full name into first and last name."""
    parts = name.strip().split()
    if len(parts) == 0:
        return "", ""
    elif len(parts) == 1:
        return parts[0], ""
    else:
        return parts[0], " ".join(parts[1:])


def main():

    wb, ws, all_rows = read_xlsx()

    to_process = [r for r in all_rows if r["processed"].lower() != "yes" and r["email"]]

    log("📋", f"Loaded {len(to_process)} entries to process "
        f"({len(all_rows)} total, skipping already-processed and missing emails)")

    found_count = 0
    created_count = 0
    note_updated_count = 0
    note_skipped_count = 0
    added_to_list_count = 0
    already_on_list_count = 0
    error_count = 0

    driver = webdriver.Chrome()
    driver.maximize_window()

    try:
        login(driver)

        for seq, row in enumerate(to_process, 1):
            email = row["email"]
            name = row["name"]
            show_role = row["show_role"]
            first, last = split_name(name)

            log("👤", f"[{seq}/{len(to_process)}] {name} ({email}) — {show_role}")

            success = False
            try:
                clear_screen(driver)
                found = lookup_by_email(driver, email)

                if found == "multiple":
                    log("⚠️", f"  MULTIPLE MATCHES — please select the correct patron in the browser")
                    input("  ⏸️  Press Enter once you've selected the patron...")
                    # Verify we're now on a patron record
                    try:
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located(
                                (By.XPATH, "//input[@value='Exit Record']")
                            )
                        )
                        log("✅", f"  Patron selected")
                        found_count += 1
                    except TimeoutException:
                        log("❌", f"  No patron record loaded — skipping")
                        error_count += 1
                        navigate_to_people_page(driver)
                        continue
                elif found:
                    log("✅", f"  Found in ArtsPeople")
                    found_count += 1
                else:
                    log("🆕", f"  Not found — creating new patron...")
                    create_patron(driver, first, last, email)
                    created_count += 1
                    # Save the new record first so the Note field becomes available
                    update_patron_record(driver)

                # Add to Volunteers list
                if is_on_volunteers_list(driver):
                    log("ℹ️", f"  Already on Volunteers list")
                    already_on_list_count += 1
                else:
                    if add_to_volunteers_list(driver):
                        log("✅", f"  Added to Volunteers list")
                        added_to_list_count += 1
                    else:
                        log("⚠️", f"  Failed to add to Volunteers list")

                # Append role to Note to House Manager
                # Replace " | " separator with em dash
                note_text = show_role.replace(" | ", " // ")
                was_added = append_note_to_house_manager(driver, note_text)
                if was_added:
                    note_updated_count += 1
                else:
                    note_skipped_count += 1

                update_patron_record(driver)
                exit_record(driver)
                success = True

            except Exception as e:
                log("❌", f"  Error processing: {e}")
                error_count += 1
                try:
                    navigate_to_people_page(driver)
                except Exception:
                    pass

            if success:
                ws.cell(row=row["row_num"], column=row["proc_col_idx"], value="Yes")
                save_xlsx(wb)

    finally:
        print("\n" + "=" * 50)
        print("  SUMMARY")
        print("=" * 50)
        print(f"  Total processed:       {found_count + created_count}")
        print(f"  Found in ArtsPeople:   {found_count}")
        print(f"  Created new patrons:   {created_count}")
        print(f"  Notes updated:         {note_updated_count}")
        print(f"  Notes already present: {note_skipped_count}")
        print(f"  Added to Volunteers:   {added_to_list_count}")
        print(f"  Already on list:       {already_on_list_count}")
        print(f"  Errors:                {error_count}")
        print("=" * 50)

        input("\nPress Enter to close the browser...")
        driver.quit()


if __name__ == "__main__":
    main()
